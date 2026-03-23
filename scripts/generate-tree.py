"""
Generate v1-tree.json from Jabari's Movement Modifications spreadsheet.

Usage:
  python scripts/generate-tree.py "path/to/Movement Modifications.xlsx"

Output:
  src/decision-tree/v1-tree.json
"""

import json
import re
import sys
import openpyxl
from pathlib import Path


# ── Mapping: spreadsheet activity name → controlled vocabulary value ──
# Activities not in VALID_ACTIVITIES will need to be added to controlled-vocabulary.ts
ACTIVITY_MAP = {
    "Split Squats": "split_squats",
    "Forward lunge": "forward_lunge",
    "Backward lunge": "backward_lunge",
    "Side lunge": "side_lunge",
    "Running (level ground)": "running_level",
    "Running (level, uneven ground)": "running_uneven",
    "Running (uphill)": "running_uphill",
    "Running (downhill)": "running_downhill",
    "Cycling": "cycling",
    "Hiking/walking (level ground)": "walking_level",
    "Hiking/walking (uphill)": "walking_uphill",
    "Hiking/walking dowmhill": "walking_downhill",
    "Squatting/deep squatting (light or no weight)": "squatting_bodyweight",
    "Squatting (heavy weight, barbell)": "squatting_barbell",
    "Ascending stairs": "stairs_up",
    "Descending stairs": "stairs_down",
    "Half kneeling (focus on overall positioning/comfort)": "half_kneeling",
    "Tall kneeling": "tall_kneeling",
    "Full kneeling": "full_kneeling",
    "Bending down (focus on teaching bending mechancis as a whole)": "bending_down",
    "Sitting down (essentially take from squatting)": "sitting_down",
    "Standing up": "standing_up",
    "Rowing machine": "rowing_machine",
    "Jumping": "jumping",
    "Cutting/pivoting": "pivoting",
    "Prolonged sitting": "prolonged_sitting",
    "Prolonged standing": "prolonged_standing",
    "Deadlifts": "deadlifts",
    "RDL": "rdl",
    "Twisting while carrying a load": "twisting_loaded",
    # Yoga
    "Hero pose": "yoga_hero",
    "Warrior pose": "yoga_warrior",
    "Eagle pose": "yoga_eagle",
    "Triangle pose (hypreextension)": "yoga_triangle",
    "Revolved chair pose": "yoga_revolved_chair",
    "Pigeon pose": "yoga_pigeon",
}

# ── Mapping: spreadsheet pain_location → controlled vocabulary location ──
LOCATION_MAP = {
    "Anteromedial femoral condyle": "anteromedial_femoral_condyle",
    "Anterolateral femoral condyle": "anterolateral_femoral_condyle",
    "Patellar tendon": "patellar_tendon",
    "Patella": "patella",
    "Patellofemoral joint (Patella)": "patella",
    "Posteromedial femoral condyle": "posteromedial_femoral_condyle",
    "Posterior femoral condyle (medial)": "posteromedial_femoral_condyle",
    "patellar/anterior knee pain; central or posterior pain": "patella",
}


def slugify(text: str) -> str:
    """Convert text to a snake_case slug for use as node/recommendation IDs."""
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9\s]", "", text)
    text = re.sub(r"\s+", "_", text)
    # Truncate to keep IDs reasonable
    return text[:60]


def extract_mod_title(mod_text: str) -> str:
    """Extract a short title from the modification text (text before the first parenthetical)."""
    # Take text before first '(' as the title/summary
    paren_idx = mod_text.find("(")
    if paren_idx > 0:
        title = mod_text[:paren_idx].strip()
    else:
        title = mod_text.strip()
    # Clean up trailing punctuation
    title = title.rstrip(".,;:")
    # Truncate if too long
    if len(title) > 120:
        title = title[:117] + "..."
    return title


def parse_spreadsheet(filepath: str) -> list[dict]:
    """Parse the Excel spreadsheet into a list of row dicts."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        activity = row[0]
        if not activity:
            continue

        pain_location_raw = row[1] or ""
        pain_location = row[2] or ""
        impairment_sources = row[3] or ""
        modifications_clinical = row[4] or ""

        # Extract modifications (6 columns per mod: text, video_id, title, steps, why, tags)
        mods = []
        MOD_STRIDE = 6
        for i in range(5, min(len(row), 47), MOD_STRIDE):
            mod_text = row[i]
            video_id = row[i + 1] if (i + 1) < len(row) else None
            title = row[i + 2] if (i + 2) < len(row) else None
            steps = row[i + 3] if (i + 3) < len(row) else None
            why = row[i + 4] if (i + 4) < len(row) else None
            tags = row[i + 5] if (i + 5) < len(row) else None

            if mod_text and str(mod_text).strip():
                mods.append({
                    "text": str(mod_text).strip(),
                    "video_id": str(video_id).strip() if video_id else None,
                    "title": str(title).strip() if title else None,
                    "steps": str(steps).strip() if steps else None,
                    "why": str(why).strip() if why else None,
                    "tags": [t.strip() for t in str(tags).split(",")] if tags else None,
                })

        rows.append({
            "activity": activity.strip(),
            "pain_location_raw": pain_location_raw.strip(),
            "pain_location": pain_location.strip(),
            "impairment_sources": impairment_sources.strip(),
            "modifications_clinical": modifications_clinical.strip(),
            "mods": mods,
        })

    return rows


def build_tree(rows: list[dict]) -> dict:
    """Build the decision tree JSON from parsed spreadsheet rows."""

    nodes = {}
    # Track activities seen and their location sub-branches
    activity_branches = {}  # activity_slug -> list of {condition, next_node_id}

    for row in rows:
        activity_raw = row["activity"]
        activity_slug = ACTIVITY_MAP.get(activity_raw)
        if not activity_slug:
            print(f"WARNING: No mapping for activity '{activity_raw}', skipping")
            continue

        location_raw = row["pain_location"]
        location_slug = LOCATION_MAP.get(location_raw)
        if not location_slug:
            print(f"WARNING: No mapping for location '{location_raw}' (activity: {activity_raw}), skipping")
            continue

        # Skip rows with no modifications
        if not row["mods"]:
            print(f"SKIP (no mods): {activity_raw} / {location_raw}")
            continue

        # ── Build assessment node for this activity+location combo ──
        assessment_id = f"dx_{activity_slug}_{location_slug}"

        # Handle duplicate assessment IDs (same activity+location, different raw pain)
        counter = 1
        unique_id = assessment_id
        while unique_id in nodes:
            counter += 1
            unique_id = f"{assessment_id}_{counter}"
        assessment_id = unique_id

        recommendations = []
        for mod_idx, mod in enumerate(row["mods"], start=1):
            rec_id = f"{assessment_id}_mod{mod_idx}"
            # Use structured title if available, fall back to extracting from text
            title = mod["title"] if mod["title"] else extract_mod_title(mod["text"])

            rec = {
                "id": rec_id,
                "title": title,
                "type": "movement_mod",
                "description": mod["text"],
            }
            if mod["video_id"]:
                rec["video_id"] = mod["video_id"]
            if mod["steps"]:
                rec["steps"] = mod["steps"]
            if mod["why"]:
                rec["why"] = mod["why"]
            if mod["tags"]:
                rec["tags"] = mod["tags"]

            recommendations.append(rec)

        nodes[assessment_id] = {
            "id": assessment_id,
            "type": "assessment",
            "label": f"{activity_raw} - {row['pain_location_raw'] or location_raw}",
            "summary": f"Assessment for {activity_raw} with {location_raw.lower()} pain",
            "explanation": row["impairment_sources"] if row["impairment_sources"] else f"Movement modification for {activity_raw}",
            "region_id": location_slug,
            "confidence": "high",
            "recommendations": recommendations,
        }

        # ── Track branch from location question to this assessment ──
        location_question_id = f"q_location_{activity_slug}"

        if location_question_id not in activity_branches:
            activity_branches[location_question_id] = {
                "activity_raw": activity_raw,
                "activity_slug": activity_slug,
                "branches": [],
                "locations_seen": set(),
            }

        # Build condition — use the location slug
        condition_key = location_slug
        # If we already have this location for this activity, use the raw pain location
        branch_info = activity_branches[location_question_id]
        if condition_key in branch_info["locations_seen"]:
            # Duplicate location — use raw pain description to differentiate
            condition_key = slugify(row["pain_location_raw"]) if row["pain_location_raw"] else f"{location_slug}_{counter}"

        branch_info["locations_seen"].add(condition_key)
        branch_info["branches"].append({
            "condition": {
                "type": "equals",
                "key": "symptom_location",
                "value": location_slug,
            },
            "next_node_id": assessment_id,
        })

    # ── Build location question nodes ──
    activity_options_branches = []  # branches for the top-level activity question

    for loc_q_id, info in activity_branches.items():
        # Collect unique location options for this activity
        seen_values = set()
        unique_branches = []
        options = []

        for branch in info["branches"]:
            value = branch["condition"]["value"]
            if value not in seen_values:
                seen_values.add(value)
                unique_branches.append(branch)
                # Human-readable label from the location slug
                label = value.replace("_", " ").title()
                options.append({"value": value, "label": label})

        # If only one location, skip the question and go directly to assessment
        if len(unique_branches) == 1:
            # Point activity directly to the assessment
            activity_options_branches.append({
                "condition": {
                    "type": "equals",
                    "key": "triggering_activity",
                    "value": info["activity_slug"],
                },
                "next_node_id": unique_branches[0]["next_node_id"],
            })
        else:
            nodes[loc_q_id] = {
                "id": loc_q_id,
                "type": "question",
                "label": f"Pain location for {info['activity_raw']}",
                "prompt": "Where exactly do you feel the pain?",
                "answer_type": "choice",
                "save_to": "symptom_location",
                "options": options,
                "next": unique_branches,
            }

            activity_options_branches.append({
                "condition": {
                    "type": "equals",
                    "key": "triggering_activity",
                    "value": info["activity_slug"],
                },
                "next_node_id": loc_q_id,
            })

    # ── Build top-level activity question ──
    activity_options = []
    for info in activity_branches.values():
        activity_options.append({
            "value": info["activity_slug"],
            "label": info["activity_raw"],
        })

    # Deduplicate activity options
    seen = set()
    unique_activity_options = []
    for opt in activity_options:
        if opt["value"] not in seen:
            seen.add(opt["value"])
            unique_activity_options.append(opt)

    nodes["q_activity"] = {
        "id": "q_activity",
        "type": "question",
        "label": "Triggering activity",
        "prompt": "Which activity triggers your knee pain the most?",
        "answer_type": "choice",
        "save_to": "triggering_activity",
        "options": unique_activity_options,
        "next": activity_options_branches,
    }

    # ── Assemble the full tree ──
    tree = {
        "id": "neez_v1",
        "version": "1.0.0",
        "title": "neez Movement Modification Decision Tree v1",
        "description": "Production decision tree generated from Jabari's Movement Modifications spreadsheet",
        "entry_node_id": "q_activity",
        "nodes": nodes,
    }

    return tree


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/generate-tree.py <path-to-spreadsheet.xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    output_path = Path(__file__).parent.parent / "src" / "decision-tree" / "v1-tree.json"

    print(f"Reading spreadsheet: {xlsx_path}")
    rows = parse_spreadsheet(xlsx_path)
    print(f"Parsed {len(rows)} rows")

    tree = build_tree(rows)

    node_count = len(tree["nodes"])
    assessment_count = sum(1 for n in tree["nodes"].values() if n["type"] == "assessment")
    question_count = sum(1 for n in tree["nodes"].values() if n["type"] == "question")
    rec_count = sum(
        len(n.get("recommendations", []))
        for n in tree["nodes"].values()
        if n["type"] == "assessment"
    )
    video_count = sum(
        1
        for n in tree["nodes"].values()
        if n["type"] == "assessment"
        for r in n.get("recommendations", [])
        if r.get("video_id")
    )

    print(f"\nTree generated:")
    print(f"  Nodes: {node_count} ({question_count} questions, {assessment_count} assessments)")
    print(f"  Recommendations: {rec_count}")
    print(f"  With video_id: {video_count}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(tree, f, indent=2, ensure_ascii=False)

    print(f"\nWritten to: {output_path}")

    # ── Print new activities that need to be added to controlled-vocabulary.ts ──
    existing_activities = [
        "running", "squatting", "lunging", "stairs_up", "stairs_down",
        "jumping", "cycling", "walking", "sitting", "kneeling", "pivoting", "other",
    ]
    new_activities = sorted(set(
        slug for slug in ACTIVITY_MAP.values()
        if slug not in existing_activities
    ))
    if new_activities:
        print(f"\nWARNING: New activities to add to controlled-vocabulary.ts:")
        for a in new_activities:
            print(f"  '{a}',")


if __name__ == "__main__":
    main()
